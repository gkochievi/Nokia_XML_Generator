"""End-to-end tests for ModernizationGenerator.generate().

Calls the generator directly (not via HTTP) so we can introspect
`replacement_counts` and read back the generated XML to confirm specific
mutations landed. These tests are the safety net for the 16-pass pipeline —
if a pass silently no-ops or order drifts, the counts assertions catch it.
"""
import os
import pytest

from modules.modernization import ModernizationGenerator


@pytest.fixture()
def generated_dir(tmp_path):
    """Standalone generated-folder fixture (not tied to the Flask app)."""
    d = tmp_path / "generated"
    d.mkdir()
    return str(d)


@pytest.fixture()
def modernization_result(rich_xml_pair, ip_plan_xlsx, generated_dir):
    """Run a full modernization generation against the rich fixtures and
    return (counts dict, output xml text, full extra dict, debug_log)."""
    existing, reference = rich_xml_pair
    ip_plan = ip_plan_xlsx('Existing_Station')

    gen = ModernizationGenerator()
    filename, debug_log, extra = gen.generate(
        station_name='Existing_Station',
        existing_xml_path=existing,
        reference_5g_xml_path=reference,
        transmission_excel_path=ip_plan,
        output_folder=generated_dir,
        existing_bts_name='Existing_Station',
        reference_bts_name='Reference_Template_5G',
        ip_plan_excel_path=ip_plan,
        mode='modernization',
    )

    output_path = os.path.join(generated_dir, filename)
    assert os.path.exists(output_path), f"output {filename} not written"
    with open(output_path, 'r', encoding='utf-8') as f:
        xml = f.read()

    counts = extra.get('replacement_counts') or {}
    return counts, xml, extra, debug_log


# ---------------------------------------------------------------------------
# Pipeline coverage — each assertion guards one of the 16 passes
# ---------------------------------------------------------------------------

def test_station_names_replaced(modernization_result):
    counts, xml, _, _ = modernization_result
    assert counts.get('station_names', 0) > 0
    # Reference name gone, target name present
    assert 'Reference_Template_5G' not in xml
    assert 'Existing_Station' in xml


def test_bts_ids_replaced(modernization_result):
    counts, xml, _, _ = modernization_result
    assert counts.get('bts_ids', 0) > 0
    assert 'MRBTS-11111' not in xml
    assert 'MRBTS-90217' in xml
    assert 'LNBTS-11111' not in xml
    assert 'LNBTS-90217' in xml


def test_vlan_ids_replaced(modernization_result):
    counts, xml, _, _ = modernization_result
    # Reference has 3 VLANIFs (OAM/LTE/NR), IP Plan supplies all three →
    # at least 3 successful replacements expected.
    assert counts.get('vlan_ids', 0) >= 3
    # Reference VLANs (3900/3940/3950) gone, IP Plan VLANs (100/400/500) present
    assert '<p name="vlanId">3900</p>' not in xml
    assert '<p name="vlanId">3940</p>' not in xml
    assert '<p name="vlanId">3950</p>' not in xml


def test_ip_addresses_replaced(modernization_result):
    counts, xml, _, _ = modernization_result
    assert counts.get('ip_addresses', 0) > 0
    # IP Plan IPs present
    assert '10.110.99.10' in xml
    assert '10.111.99.10' in xml
    assert '10.112.99.10' in xml
    # Reference per-tech IPs gone
    assert '10.110.0.10' not in xml
    assert '10.111.0.10' not in xml
    assert '10.112.0.10' not in xml


def test_gateways_replaced(modernization_result):
    counts, xml, _, _ = modernization_result
    assert counts.get('gateways', 0) >= 1
    # IPRT-2 (5G) and IPRT-1 (OAM dest=0.0.0.0 + 4G dest=10.111.0.0) should be swapped
    assert '10.112.99.1' in xml  # 5G GW from IP Plan, written into IPRT-2


def test_network_params_structural(modernization_result):
    counts, xml, _, _ = modernization_result
    # NRX2LINK_TRUST-1 → LTE IP, LNADJGNB-0 → 5G IP
    assert counts.get('network_params_structural', 0) >= 2
    assert '10.111.99.10' in xml  # LTE IP landed in NRX2LINK_TRUST-1
    assert '10.112.99.10' in xml  # 5G IP landed in LNADJGNB-0


def test_4g_cells_replaced(modernization_result):
    """_replace_4g_cells writes phyCellId and tac from existing into reference."""
    counts, xml, _, _ = modernization_result
    assert counts.get('cells_4g', 0) > 0
    # Reference LNCEL-11 had phyCellId=500, tac=9999
    # Existing had phyCellId=100, tac=5001
    assert '<p name="phyCellId">500</p>' not in xml
    assert '<p name="phyCellId">100</p>' in xml


def test_4g_rootseq_replaced(modernization_result):
    """_replace_4g_rootseq writes rootSeqIndex from existing LNCEL_FDD into
    reference LNCEL_FDD. This pass is separate from _replace_4g_cells because
    rootSeqIndex lives on the LNCEL_FDD child, not the LNCEL parent — and
    extract_4g_cells does not extract it."""
    counts, xml, _, _ = modernization_result
    assert counts.get('rootseq_4g', 0) > 0
    # Reference LNCEL_FDD-0 had rootSeqIndex=10; existing had 620
    assert '<p name="rootSeqIndex">10</p>' not in xml
    assert '<p name="rootSeqIndex">620</p>' in xml


def test_5g_nrcell_physcellid_mapped_from_4g(modernization_result):
    counts, xml, _, _ = modernization_result
    assert counts.get('nrcells_5g_pci', 0) > 0
    # NRCELL-111 → maps to LNCEL-11 (last 2 digits) → takes existing's phyCellId=100
    assert '<p name="physCellId">900</p>' not in xml
    assert '<p name="physCellId">100</p>' in xml


def test_iot_tac_forced_to_5000(modernization_result):
    """LNCEL-211..214 must always have TAC=5000, even though the reference
    template had TAC=7777 for LNCEL-211 and the rollout TAC override is not
    set in modernization mode."""
    counts, xml, _, _ = modernization_result
    assert counts.get('iot_tac_fix', 0) >= 1
    # The reference's LNCEL-211 block had tac=7777 — must now be 5000
    assert '<p name="tac">7777</p>' not in xml
    # Find the LNCEL-211 block specifically and assert TAC=5000 is inside it
    assert 'LNCEL-211' in xml
    iot_start = xml.find('LNCEL-211')
    iot_end = xml.find('</managedObject>', iot_start)
    assert iot_start >= 0 and iot_end > iot_start
    iot_block = xml[iot_start:iot_end]
    assert '<p name="tac">5000</p>' in iot_block


def test_ip_plan_found_flag_true(modernization_result):
    _, _, extra, _ = modernization_result
    assert extra.get('ip_plan_found') is True
    assert extra.get('ip_plan_lookup') == 'Existing_Station'


# ---------------------------------------------------------------------------
# Specific guards for fixed bugs from the May 2026 refactor
# ---------------------------------------------------------------------------

def test_short_reference_name_does_not_substring_match(
    tmp_path, rich_xml_pair, ip_plan_xlsx, generated_dir
):
    """_replace_station_names refuses to operate when the reference name is
    shorter than MIN_NAME_TOKEN_LEN. Without this guard, e.g. ref name 'AB'
    would match the 'AB' inside 'NRBTS-90217' and other tokens."""
    existing, _ = rich_xml_pair

    # Build a reference whose btsName is a too-short token
    short_ref = tmp_path / "ref_short.xml"
    short_ref.write_text(
        '<?xml version="1.0"?>\n'
        '<raml xmlns="raml21.xsd"><cmData>'
        '<managedObject class="com.nokia.srbts:MRBTS" distName="MRBTS-1">'
        '<p name="btsName">AB</p>'
        '</managedObject>'
        '</cmData></raml>',
        encoding='utf-8',
    )
    ip_plan = ip_plan_xlsx('Existing_Station')

    gen = ModernizationGenerator()
    _, _, extra = gen.generate(
        station_name='Existing_Station',
        existing_xml_path=existing,
        reference_5g_xml_path=str(short_ref),
        transmission_excel_path=ip_plan,
        output_folder=generated_dir,
        existing_bts_name='Existing_Station',
        reference_bts_name='AB',
        ip_plan_excel_path=ip_plan,
        mode='modernization',
    )
    counts = extra.get('replacement_counts') or {}
    assert counts.get('station_names', -1) == 0, \
        "short reference name must not trigger any substring replacement"


def test_ip_plan_lookup_tolerates_whitespace(
    tmp_path, rich_xml_pair, generated_dir
):
    """Station name match collapses whitespace, so a trailing space or
    non-breaking space in the Excel cell should still be found."""
    import openpyxl
    from constants import IP_PLAN_COLUMNS

    existing, reference = rich_xml_pair
    plan_path = tmp_path / "ip_plan_messy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    # Insert station with trailing space + a non-breaking space mid-token
    ws.cell(row=2, column=1, value='Existing Station ')
    values = {
        'MGT_VLAN_ID': 100,  'MGT_IP': '10.110.99.10',  'MGT_MASK': 24, 'MGT_GW': '10.110.99.1',
        'GSM_VLAN_ID': 200,  'GSM_IP': '10.171.99.10',  'GSM_MASK': 24, 'GSM_GW': '10.171.99.1',
        'WCDMA_VLAN_ID': 300,'WCDMA_IP':'10.141.99.10', 'WCDMA_MASK':24,'WCDMA_GW':'10.141.99.1',
        'LTE_VLAN': 400,     'LTE_IP': '10.111.99.10',  'LTE_MASK': 24, 'LTE_GW': '10.111.99.1',
        '5G_VLAN': 500,      '5G_IP':  '10.112.99.10',  '5G_MASK':  24, '5G_GW':  '10.112.99.1',
    }
    for field, col_idx in IP_PLAN_COLUMNS.items():
        ws.cell(row=2, column=col_idx + 1, value=values[field])
    wb.save(str(plan_path))

    gen = ModernizationGenerator()
    _, _, extra = gen.generate(
        station_name='ExistingStation',  # no separator, will match via whitespace collapse
        existing_xml_path=existing,
        reference_5g_xml_path=reference,
        transmission_excel_path=str(plan_path),
        output_folder=generated_dir,
        existing_bts_name='Existing_Station',
        reference_bts_name='Reference_Template_5G',
        ip_plan_excel_path=str(plan_path),
        mode='modernization',
    )
    assert extra.get('ip_plan_found') is True


def test_rollout_mode_applies_tac_override(
    rich_xml_pair, ip_plan_xlsx, generated_dir
):
    """Rollout mode used to silently drop the TAC override (the /api/rollout
    route passed tac=None). Now both routes honor it. Test the generator
    contract directly."""
    _, reference = rich_xml_pair
    ip_plan = ip_plan_xlsx('Rollout_Station')

    gen = ModernizationGenerator()
    _, _, extra = gen.generate(
        station_name='Rollout_Station',
        existing_xml_path=reference,  # rollout reuses reference as existing
        reference_5g_xml_path=reference,
        transmission_excel_path=ip_plan,
        output_folder=generated_dir,
        existing_bts_name='Reference_Template_5G',
        reference_bts_name='Reference_Template_5G',
        ip_plan_excel_path=ip_plan,
        mode='rollout',
        rollout_overrides={'id': '88888', 'name': 'Rollout_Station', 'tac': '4242'},
    )
    counts = extra.get('replacement_counts') or {}
    # _override_tac_all visits every LNCEL (including IoT) — must touch at
    # least the two LNCELs in the reference (LNCEL-11 and LNCEL-211).
    assert counts.get('tac_override', 0) >= 2
    # And _fix_iot_tac must STILL run last and force 5000 on LNCEL-211 even
    # though the rollout TAC override was 4242.
    assert counts.get('iot_tac_fix', 0) >= 1


# ---------------------------------------------------------------------------
# Post-generation verification (_verify_output)
# ---------------------------------------------------------------------------

def test_verification_passes_on_clean_run(modernization_result):
    """A normal modernization against the rich fixtures must produce no
    verification errors and no warnings."""
    _, _, extra, _ = modernization_result
    verification = extra.get('verification') or {}
    assert verification.get('errors') == [], \
        f"clean run produced verification errors: {verification.get('errors')}"
    assert verification.get('warnings') == [], \
        f"clean run produced verification warnings: {verification.get('warnings')}"


def test_verification_catches_unreplaced_reference_name(
    tmp_path, rich_xml_pair, ip_plan_xlsx, generated_dir
):
    """If the reference btsName is too short for safe substring replacement
    (<MIN_NAME_TOKEN_LEN), _replace_station_names skips. The resulting output
    still contains the reference name and is unsafe to deploy — verification
    must catch this."""
    from conftest import REFERENCE_TEMPLATE_XML

    existing, _ = rich_xml_pair
    # Build a reference XML where btsName is "AB" but the rest of the file is
    # large enough to pass the size check.
    corrupted_ref = tmp_path / "ref_short_name.xml"
    corrupted_ref.write_text(
        REFERENCE_TEMPLATE_XML
            .replace('Reference_Template_5G', 'AB')
            .replace('<p name="nrBtsName">AB</p>', '<p name="nrBtsName">AB</p>'),
        encoding='utf-8',
    )
    ip_plan = ip_plan_xlsx('Existing_Station')

    gen = ModernizationGenerator()
    _, _, extra = gen.generate(
        station_name='Existing_Station',
        existing_xml_path=existing,
        reference_5g_xml_path=str(corrupted_ref),
        transmission_excel_path=ip_plan,
        output_folder=generated_dir,
        existing_bts_name='Existing_Station',
        reference_bts_name='AB',
        ip_plan_excel_path=ip_plan,
        mode='modernization',
    )
    verification = extra.get('verification') or {}
    errors = verification.get('errors') or []
    assert errors, "verification must fail when reference btsName is left in output"
    assert any("btsName 'AB' still present" in e for e in errors), \
        f"expected btsName-still-present error, got: {errors}"


def test_verification_skips_ref_equality_in_rollout_mode(
    rich_xml_pair, ip_plan_xlsx, generated_dir
):
    """In rollout mode, the reference XML is used as both existing and reference,
    so reference_bts_name == target_name by design. The ref-equality checks must
    NOT fire — that would falsely block every rollout."""
    _, reference = rich_xml_pair
    ip_plan = ip_plan_xlsx('Rollout_Station')

    gen = ModernizationGenerator()
    _, _, extra = gen.generate(
        station_name='Rollout_Station',
        existing_xml_path=reference,
        reference_5g_xml_path=reference,
        transmission_excel_path=ip_plan,
        output_folder=generated_dir,
        existing_bts_name='Reference_Template_5G',
        reference_bts_name='Reference_Template_5G',
        ip_plan_excel_path=ip_plan,
        mode='rollout',
        rollout_overrides={'id': '88888', 'name': 'Rollout_Station', 'tac': '4242'},
    )
    verification = extra.get('verification') or {}
    errors = verification.get('errors') or []
    # ref name and ref id checks are skipped when source==target. No errors expected.
    assert not any("btsName" in e for e in errors), \
        f"rollout mode incorrectly flagged btsName equality: {errors}"
    assert not any("BTS ID still present" in e for e in errors), \
        f"rollout mode incorrectly flagged BTS ID equality: {errors}"


def test_verification_logs_to_debug_log(modernization_result):
    """A clean pass should append a ✓ [VERIFY] success line to debug_log so
    the operator sees the check ran."""
    _, _, _, debug_log = modernization_result
    assert any('[VERIFY]' in line for line in debug_log), \
        "debug_log must include a [VERIFY] entry"


def test_replacement_counts_keys_are_stable(modernization_result):
    """Lock the set of replacement_counts keys so a renamed counter shows up
    as a test failure rather than a silent frontend break."""
    counts, _, _, _ = modernization_result
    expected = {
        'station_names', 'bts_ids', 'vlan_ids', 'ip_addresses', 'gateways',
        'network_params_structural', 'network_params_legacy',
        'sctp_port_min', 'params_2g', 'cells_4g', 'rootseq_4g',
        'nrcells_5g_pci', 'tdd_cells_4g', 'tdd_pci_from_fdd',
        'nrcell_5g_details', 'tac_override', 'iot_tac_fix',
    }
    # All counters that ran should be in the expected set. Some won't fire
    # (e.g. params_2g — no 2G in the rich fixture); that's fine. Just ensure
    # nothing UNEXPECTED is present.
    unexpected = set(counts.keys()) - expected
    assert not unexpected, f"unexpected replacement_counts keys: {unexpected}"
