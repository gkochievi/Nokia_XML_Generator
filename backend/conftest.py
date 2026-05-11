import os
import shutil
import tempfile
import pytest


@pytest.fixture()
def app():
    """Create a Flask application configured for testing."""
    # Create temp dirs so tests don't touch real data
    tmp = tempfile.mkdtemp()
    upload_dir = os.path.join(tmp, 'uploads')
    generated_dir = os.path.join(tmp, 'generated')
    example_dir = os.path.join(tmp, 'example_files')
    os.makedirs(upload_dir)
    os.makedirs(generated_dir)
    os.makedirs(example_dir)
    os.makedirs(os.path.join(example_dir, 'East'))
    os.makedirs(os.path.join(example_dir, 'West'))
    os.makedirs(os.path.join(example_dir, 'IP'))

    from app import app as flask_app
    flask_app.config.update({
        'TESTING': True,
        'UPLOAD_FOLDER': upload_dir,
        'GENERATED_FOLDER': generated_dir,
        'EXAMPLE_FILES_FOLDER': example_dir,
    })

    yield flask_app

    shutil.rmtree(tmp, ignore_errors=True)


@pytest.fixture()
def client(app):
    """Flask test client."""
    return app.test_client()


@pytest.fixture()
def example_dir(app):
    """Path to the test example_files directory."""
    return app.config['EXAMPLE_FILES_FOLDER']


@pytest.fixture()
def generated_dir(app):
    """Path to the test generated directory."""
    return app.config['GENERATED_FOLDER']


# ---------------------------------------------------------------------------
# Minimal Nokia XML fixtures (in-memory strings)
# ---------------------------------------------------------------------------

MINIMAL_NOKIA_XML = """\
<?xml version="1.0" encoding="UTF-8"?>
<raml xmlns="raml21.xsd" version="2.1">
  <cmData type="plan" scope="all" id="1234">
    <header>
      <log action="create" dateTime="2026-01-01T00:00:00Z"/>
    </header>
    <managedObject class="com.nokia.srbts:MRBTS" distName="MRBTS-12345" operation="create">
      <p name="btsName">Test_Station_Alpha</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:WNBTS" distName="MRBTS-12345/WNBTS-12345" operation="create">
      <list name="cPlaneList">
        <item>
          <p name="sctpPortMin">2905</p>
        </item>
      </list>
    </managedObject>
    <managedObject class="NOKLTE:LNBTS" distName="MRBTS-12345/LNBTS-12345" operation="create">
      <p name="lnBtsId">12345</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL" distName="MRBTS-12345/LNBTS-12345/LNCEL-11" operation="create">
      <p name="phyCellId">100</p>
      <p name="tac">5001</p>
      <p name="eutraCelId">11</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL" distName="MRBTS-12345/LNBTS-12345/LNCEL-12" operation="create">
      <p name="phyCellId">101</p>
      <p name="tac">5001</p>
      <p name="eutraCelId">12</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL_FDD" distName="MRBTS-12345/LNBTS-12345/LNCEL-11/LNCEL_FDD-11" operation="create">
      <p name="rootSeqIndex">620</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL_FDD" distName="MRBTS-12345/LNBTS-12345/LNCEL-12/LNCEL_FDD-12" operation="create">
      <p name="rootSeqIndex">350</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRBTS" distName="MRBTS-12345/NRBTS-12345" operation="create">
      <p name="nrBtsName">Test_Station_Alpha</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRCELL" distName="MRBTS-12345/NRBTS-12345/NRCELL-111" operation="create">
      <p name="physCellId">200</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRCELL" distName="MRBTS-12345/NRBTS-12345/NRCELL-112" operation="create">
      <p name="physCellId">201</p>
    </managedObject>
  </cmData>
</raml>
"""

MINIMAL_NOKIA_XML_NO_NAMESPACE = """\
<?xml version="1.0" encoding="UTF-8"?>
<raml version="2.1">
  <cmData type="plan" scope="all" id="9999">
    <header>
      <log action="create" dateTime="2026-01-01T00:00:00Z"/>
    </header>
    <managedObject class="com.nokia.srbts:MRBTS" distName="MRBTS-99999" operation="create">
      <p name="btsName">NoNS_Station</p>
    </managedObject>
  </cmData>
</raml>
"""


@pytest.fixture()
def sample_xml_content():
    """Return minimal Nokia XML string."""
    return MINIMAL_NOKIA_XML


@pytest.fixture()
def sample_xml_file(tmp_path):
    """Write minimal Nokia XML to a temp file, return path."""
    p = tmp_path / "test_station.xml"
    p.write_text(MINIMAL_NOKIA_XML, encoding='utf-8')
    return str(p)


@pytest.fixture()
def sample_xml_no_ns_file(tmp_path):
    """Write Nokia XML without default namespace to a temp file."""
    p = tmp_path / "test_no_ns.xml"
    p.write_text(MINIMAL_NOKIA_XML_NO_NAMESPACE, encoding='utf-8')
    return str(p)


# ---------------------------------------------------------------------------
# Rich XML fixtures for end-to-end ModernizationGenerator tests
# ---------------------------------------------------------------------------
#
# These two fixtures (existing + reference) carry every managedObject family
# that the 16-pass pipeline mutates: btsName, BTS IDs, VLANIF, IPIF/IPADDRESSV4,
# IPRT staticRoutes, NRX2LINK_TRUST-1, LNADJGNB-0, LNCEL (incl. IoT 211), LNCEL_FDD
# with rootSeqIndex, NRCELL with physCellId. Reference values differ from the
# "existing" station so each replacement is observable in the output XML and in
# `replacement_counts`.

EXISTING_STATION_XML = """\
<?xml version="1.0" encoding="UTF-8"?>
<raml xmlns="raml21.xsd" version="2.1">
  <cmData type="plan" scope="all" id="1">
    <header><log action="create" dateTime="2026-01-01T00:00:00Z"/></header>
    <managedObject class="com.nokia.srbts:MRBTS" distName="MRBTS-90217">
      <p name="btsName">Existing_Station</p>
    </managedObject>
    <managedObject class="NOKLTE:LNBTS" distName="MRBTS-90217/LNBTS-90217">
      <p name="lnBtsId">90217</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL" distName="MRBTS-90217/LNBTS-90217/LNCEL-11">
      <p name="phyCellId">100</p>
      <p name="tac">5001</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL_FDD" distName="MRBTS-90217/LNBTS-90217/LNCEL-11/LNCEL_FDD-0">
      <p name="rootSeqIndex">620</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRBTS" distName="MRBTS-90217/NRBTS-90217">
      <p name="nrBtsName">Existing_Station</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRCELL" distName="MRBTS-90217/NRBTS-90217/NRCELL-111">
      <p name="physCellId">200</p>
    </managedObject>
  </cmData>
</raml>
"""


REFERENCE_TEMPLATE_XML = """\
<?xml version="1.0" encoding="UTF-8"?>
<raml xmlns="raml21.xsd" version="2.1">
  <cmData type="plan" scope="all" id="1">
    <header><log action="create" dateTime="2026-01-01T00:00:00Z"/></header>
    <managedObject class="com.nokia.srbts:MRBTS" distName="MRBTS-11111">
      <p name="btsName">Reference_Template_5G</p>
    </managedObject>
    <managedObject class="NOKLTE:LNBTS" distName="MRBTS-11111/LNBTS-11111">
      <p name="lnBtsId">11111</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:VLANIF" distName="MRBTS-11111/VLANIF-1">
      <p name="userLabel">OAM</p>
      <p name="vlanId">3900</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:VLANIF" distName="MRBTS-11111/VLANIF-2">
      <p name="userLabel">LTE</p>
      <p name="vlanId">3940</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:VLANIF" distName="MRBTS-11111/VLANIF-3">
      <p name="userLabel">NR</p>
      <p name="vlanId">3950</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPIF" distName="MRBTS-11111/IPIF-1">
      <p name="userLabel">OAM</p>
      <p name="interfaceDN">MRBTS-11111/VLANIF-1</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPADDRESSV4" distName="MRBTS-11111/IPIF-1/IPADDRESSV4-0">
      <p name="localIpAddr">10.110.0.10</p>
      <p name="localIpPrefixLength">24</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPIF" distName="MRBTS-11111/IPIF-2">
      <p name="userLabel">LTE</p>
      <p name="interfaceDN">MRBTS-11111/VLANIF-2</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPADDRESSV4" distName="MRBTS-11111/IPIF-2/IPADDRESSV4-0">
      <p name="localIpAddr">10.111.0.10</p>
      <p name="localIpPrefixLength">24</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPIF" distName="MRBTS-11111/IPIF-3">
      <p name="userLabel">NR</p>
      <p name="interfaceDN">MRBTS-11111/VLANIF-3</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPADDRESSV4" distName="MRBTS-11111/IPIF-3/IPADDRESSV4-0">
      <p name="localIpAddr">10.112.0.10</p>
      <p name="localIpPrefixLength">24</p>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPRT" distName="MRBTS-11111/IPRT-1">
      <list name="staticRoutes">
        <item><p name="destIpAddr">0.0.0.0</p><p name="gateway">10.110.0.1</p></item>
        <item><p name="destIpAddr">10.111.0.0</p><p name="gateway">10.111.0.1</p></item>
      </list>
    </managedObject>
    <managedObject class="com.nokia.srbts:IPRT" distName="MRBTS-11111/IPRT-2">
      <p name="userLabel">NR</p>
      <list name="staticRoutes">
        <item><p name="destIpAddr">10.112.0.0</p><p name="gateway">10.112.0.1</p></item>
      </list>
    </managedObject>
    <managedObject class="NOKLTE:NRX2LINK_TRUST" distName="MRBTS-11111/NRX2LINK_TRUST-1">
      <p name="ipV4Addr">10.111.0.99</p>
    </managedObject>
    <managedObject class="NOKLTE:LNADJGNB" distName="MRBTS-11111/LNADJGNB-0">
      <p name="cPlaneIpAddr">10.112.0.99</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL" distName="MRBTS-11111/LNBTS-11111/LNCEL-11">
      <p name="phyCellId">500</p>
      <p name="tac">9999</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL_FDD" distName="MRBTS-11111/LNBTS-11111/LNCEL-11/LNCEL_FDD-0">
      <p name="rootSeqIndex">10</p>
    </managedObject>
    <managedObject class="NOKLTE:LNCEL" distName="MRBTS-11111/LNBTS-11111/LNCEL-211">
      <p name="phyCellId">600</p>
      <p name="tac">7777</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRBTS" distName="MRBTS-11111/NRBTS-11111">
      <p name="nrBtsName">Reference_Template_5G</p>
    </managedObject>
    <managedObject class="com.nokia.srbts.nrbts:NRCELL" distName="MRBTS-11111/NRBTS-11111/NRCELL-111">
      <p name="physCellId">900</p>
    </managedObject>
  </cmData>
</raml>
"""


def _build_ip_plan_workbook(path, station_name):
    """Write a minimal IP Plan xlsx that matches IP_PLAN_COLUMNS layout.

    Row 1 holds the station name in column A and concrete per-tech values
    at the column indices declared in constants.IP_PLAN_COLUMNS.
    """
    import openpyxl
    from constants import IP_PLAN_COLUMNS

    wb = openpyxl.Workbook()
    ws = wb.active

    # Test values keyed by IP_PLAN_COLUMNS field name. Choose IPs that exist in
    # DEST_IP_TO_TECH so _replace_gateways_by_tech has work to do.
    values = {
        'MGT_VLAN_ID': 100,  'MGT_IP': '10.110.99.10',  'MGT_MASK': 24, 'MGT_GW': '10.110.99.1',
        'GSM_VLAN_ID': 200,  'GSM_IP': '10.171.99.10',  'GSM_MASK': 24, 'GSM_GW': '10.171.99.1',
        'WCDMA_VLAN_ID': 300,'WCDMA_IP':'10.141.99.10', 'WCDMA_MASK':24,'WCDMA_GW':'10.141.99.1',
        'LTE_VLAN': 400,     'LTE_IP': '10.111.99.10',  'LTE_MASK': 24, 'LTE_GW': '10.111.99.1',
        '5G_VLAN': 500,      '5G_IP':  '10.112.99.10',  '5G_MASK':  24, '5G_GW':  '10.112.99.1',
    }

    # Station name goes into column 0 of row 2 (1-indexed). ExcelParser scans
    # every cell, so position doesn't matter for the lookup — but the row index
    # is what's then used to read columns 6..39.
    ws.cell(row=2, column=1, value=station_name)
    for field, col_idx in IP_PLAN_COLUMNS.items():
        ws.cell(row=2, column=col_idx + 1, value=values[field])

    wb.save(path)


@pytest.fixture()
def rich_xml_pair(tmp_path):
    """Write existing+reference XMLs and return their paths.

    Use these with ModernizationGenerator.generate() to exercise the full
    replacement pipeline.
    """
    existing = tmp_path / "existing.xml"
    reference = tmp_path / "reference.xml"
    existing.write_text(EXISTING_STATION_XML, encoding='utf-8')
    reference.write_text(REFERENCE_TEMPLATE_XML, encoding='utf-8')
    return str(existing), str(reference)


@pytest.fixture()
def ip_plan_xlsx(tmp_path):
    """Factory: build an IP Plan workbook keyed on a given station name."""
    def _factory(station_name='Existing_Station'):
        path = tmp_path / f"ip_plan_{station_name}.xlsx"
        _build_ip_plan_workbook(str(path), station_name)
        return str(path)
    return _factory
